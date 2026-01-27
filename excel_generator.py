import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app import db

def generate_requests_excel(requests=None):
    """Generate Excel file with all acquisition requests"""
    from models import AcquisitionRequest, User, StatusChange, Attachment
    
    if requests is None:
        requests = AcquisitionRequest.query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos de Aquisição"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    headers = [
        "ID", "Título", "Descrição", "Status", "Classe", "Categoria",
        "Responsável", "Valor Estimado", "Valor Final", "Data Solicitação",
        "Observações", "Data Criação", "Última Atualização", "Criado por", "Anexos"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        
        # Set column width
        if header in ["Título", "Descrição"]:
            ws.column_dimensions[get_column_letter(col)].width = 30
        elif header in ["Observações", "Responsável"]:
            ws.column_dimensions[get_column_letter(col)].width = 25
        else:
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Data rows
    for row, request in enumerate(requests, 2):
        # Get responsible user name
        responsible_name = ""
        if request.responsible_id:
            responsible = User.query.get(request.responsible_id)
            if responsible:
                responsible_name = responsible.full_name
        
        # Get status display
        status_display = request.get_status_display()
        
        # Get attachments count
        attachments_count = request.attachments.count()
        attachments_text = f"{attachments_count} arquivo(s)" if attachments_count > 0 else "Nenhum anexo"
        
        # Get classe and categoria display names
        classe_display = ""
        if hasattr(request, 'classe') and request.classe:
            classe_names = {'ensino': 'Ensino', 'manutencao': 'Manutenção', 'administrativo': 'Administrativo'}
            classe_display = classe_names.get(request.classe, request.classe)
            
        categoria_display = ""
        if hasattr(request, 'categoria') and request.categoria:
            categoria_names = {'material': 'Material', 'servico': 'Serviço'}
            categoria_display = categoria_names.get(request.categoria, request.categoria)
            
        # Data
        data = [
            request.id,
            request.title,
            request.description or "",
            status_display,
            classe_display,
            categoria_display,
            responsible_name,
            f"R$ {request.estimated_value:.2f}" if request.estimated_value else "Não informado",
            f"R$ {request.final_value:.2f}" if request.final_value else "Não informado",
            request.request_date.strftime("%d/%m/%Y") if request.request_date else "",
            request.observations or "",
            request.created_at.strftime("%d/%m/%Y %H:%M") if request.created_at else "",
            request.updated_at.strftime("%d/%m/%Y %H:%M") if request.updated_at else "",
            request.creator.full_name if request.creator else "",
            attachments_text
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Color coding based on status (column 4)
            if col == 4:  # Status column
                if request.status == 'finalizado':
                    cell.fill = PatternFill(start_color="D4E6B7", end_color="D4E6B7", fill_type="solid")
                elif request.status == 'a_caminho':
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif request.status == 'fase_compra':
                    cell.fill = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
                else:  # orcamento
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Color coding for Classe (column 5)
            elif col == 5:
                if hasattr(request, 'classe'):
                    if request.classe == 'ensino':
                        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Light blue
                    elif request.classe == 'manutencao':
                        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Light lavender
                    elif request.classe == 'administrativo':
                        cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Lemon Chiffon
                        
            # Color coding for Categoria (column 6)  
            elif col == 6:
                if hasattr(request, 'categoria'):
                    if request.categoria == 'material':
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Light green
                    elif request.categoria == 'servico':
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Light yellow
    
    # Add summary sheet
    ws2 = wb.create_sheet("Resumo")
    
    # Summary headers
    ws2.cell(row=1, column=1, value="Escola SENAI \"Morvan Figueiredo\"").font = Font(bold=True, size=16)
    ws2.cell(row=2, column=1, value="Relatório de Pedidos de Aquisição").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Statistics
    total_requests = len(requests)
    status_counts = {}
    for status_code, status_name in AcquisitionRequest.STATUS_CHOICES:
        count = sum(1 for req in requests if req.status == status_code)
        status_counts[status_name] = count
    
    ws2.cell(row=5, column=1, value="Estatísticas:").font = Font(bold=True)
    ws2.cell(row=6, column=1, value=f"Total de Pedidos: {total_requests}")
    
    row_num = 7
    for status_name, count in status_counts.items():
        ws2.cell(row=row_num, column=1, value=f"{status_name}: {count}")
        row_num += 1
    
    return wb

def generate_request_excel(request_id):
    """Generate Excel file for a specific request with detailed information"""
    from models import AcquisitionRequest, User, StatusChange, Attachment
    
    request = AcquisitionRequest.query.get_or_404(request_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalhes do Pedido"
    
    # Title
    ws.cell(row=1, column=1, value="Escola SENAI \"Morvan Figueiredo\"").font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value="Detalhes do Pedido de Aquisição").font = Font(bold=True, size=14)
    
    # Request details
    ws.cell(row=4, column=1, value="ID do Pedido:").font = Font(bold=True)
    ws.cell(row=4, column=2, value=request.id)
    
    ws.cell(row=5, column=1, value="Título:").font = Font(bold=True)
    ws.cell(row=5, column=2, value=request.title)
    
    ws.cell(row=6, column=1, value="Descrição:").font = Font(bold=True)
    ws.cell(row=6, column=2, value=request.description)
    ws.cell(row=6, column=2).alignment = Alignment(wrap_text=True)
    
    ws.cell(row=7, column=1, value="Status:").font = Font(bold=True)
    ws.cell(row=7, column=2, value=request.get_status_display())
    
    ws.cell(row=8, column=1, value="Criado por:").font = Font(bold=True)
    ws.cell(row=8, column=2, value=request.creator.full_name if request.creator else "")
    
    ws.cell(row=9, column=1, value="Responsável:").font = Font(bold=True)
    responsible = User.query.get(request.responsible_id) if request.responsible_id else None
    ws.cell(row=9, column=2, value=responsible.full_name if responsible else "Não definido")
    
    ws.cell(row=10, column=1, value="Data de Criação:").font = Font(bold=True)
    ws.cell(row=10, column=2, value=request.created_at.strftime("%d/%m/%Y %H:%M") if request.created_at else "")
    
    ws.cell(row=11, column=1, value="Valor Estimado:").font = Font(bold=True)
    ws.cell(row=11, column=2, value=f"R$ {request.estimated_value:.2f}" if request.estimated_value else "Não informado")
    
    ws.cell(row=12, column=1, value="Valor Final:").font = Font(bold=True)
    ws.cell(row=12, column=2, value=f"R$ {request.final_value:.2f}" if request.final_value else "Não informado")
    
    ws.cell(row=13, column=1, value="Observações:").font = Font(bold=True)
    ws.cell(row=13, column=2, value=request.observations or "Nenhuma observação")
    ws.cell(row=13, column=2).alignment = Alignment(wrap_text=True)
    
    # Status history
    status_history = StatusChange.query.filter_by(request_id=request.id).order_by(StatusChange.change_date.desc()).all()
    if status_history:
        ws.cell(row=15, column=1, value="Histórico de Status:").font = Font(bold=True)
        
        # Headers for status history
        headers = ["Data", "Status Anterior", "Novo Status", "Alterado por", "Comentários"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=16, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row_num, change in enumerate(status_history, 17):
            ws.cell(row=row_num, column=1, value=change.change_date.strftime("%d/%m/%Y %H:%M"))
            ws.cell(row=row_num, column=2, value=change.old_status or "Início")
            ws.cell(row=row_num, column=3, value=change.new_status)
            ws.cell(row=row_num, column=4, value=change.changed_by_user.full_name if change.changed_by_user else "")
            ws.cell(row=row_num, column=5, value=change.comments or "")
    
    # Attachments
    attachments = request.attachments.all()
    if attachments:
        current_row = len(status_history) + 19 if status_history else 17
        ws.cell(row=current_row, column=1, value="Anexos:").font = Font(bold=True)
        
        for i, attachment in enumerate(attachments, current_row + 1):
            ws.cell(row=i, column=1, value=f"• {attachment.original_filename}")
            ws.cell(row=i, column=2, value=f"Enviado por: {attachment.uploaded_by.full_name}")
            ws.cell(row=i, column=3, value=f"Data: {attachment.upload_date.strftime('%d/%m/%Y %H:%M')}")
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    
    return wb