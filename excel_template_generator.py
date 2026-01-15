from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import User, AcquisitionRequest
import io

def generate_import_template():
    """Gera o modelo Excel para importação em lote de pedidos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo de Importação"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    headers = [
        "Título*", "Descrição*", "Status (Orçamento, Fase de Compra, etc)*", "Prioridade (Baixa, Media, Alta, Urgente)*", "Impacto (Baixo, Medio, Alto)*", "Classe (Ensino ou Manutenção)*", "Categoria (Material, Serviço)*", "Data (DD/MM/AAAA)*",
        "Valor Estimado", "Valor Final", "Responsável (Nome Completo)", "Observações"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        
        # Set column width
        if header in ["Título*", "Descrição*"]:
            ws.column_dimensions[get_column_letter(col)].width = 40
        elif header == "Observações":
            ws.column_dimensions[get_column_letter(col)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Add example rows
    example_data = [
        [
            "Compra de Material de Escritório",
            "Aquisição de canetas, papéis e materiais básicos para o escritório administrativo",
            "orcamento",
            "media",
            "baixo",
            "ensino",
            "material",
            "2025-08-20",
            "150.00",
            "",
            "Edson Lemes",
            "Urgente para início do semestre"
        ],
        [
            "Equipamento de Informática",
            "Computadores para laboratório de informática - 10 unidades",
            "fase_compra",
            "alta",
            "medio",
            "ensino",
            "material",
            "2025-08-15",
            "25000.00",
            "24500.00",
            "Edson Lemes",
            ""
        ],
        [
            "Serviço de Manutenção e Material",
            "Reparo do sistema elétrico com fornecimento de materiais necessários",
            "orcamento",
            "urgente",
            "alto",
            "manutencao",
            "material,servico",
            "2025-08-18",
            "800.00",
            "",
            "Edilson Costa",
            "Categoria múltipla: Material e Serviço"
        ]
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Light background for examples
            cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instruções")
    
    instructions = [
        ["INSTRUÇÕES PARA IMPORTAÇÃO EM LOTE", ""],
        ["", ""],
        ["1. Campos obrigatórios (marcados com *):", ""],
        ["   • Título: Nome do pedido (ex: Compra de Laptops)", ""],
        ["   • Descrição: Detalhes do pedido (ex: 10 unidades Dell Latitude)", ""],
        ["   • Status: Orçamento, Fase de Compra, A Caminho ou Finalizado", ""],
        ["   • Prioridade: Baixa, Media, Alta ou Urgente", ""],
        ["   • Impacto: Baixo, Medio ou Alto", ""],
        ["   • Classe: Ensino ou Manutenção", ""],
        ["   • Categoria: Material, Serviço ou Material,Serviço", ""],
        ["   • Data: Formato brasileiro (DD/MM/AAAA) ou AAAA-MM-DD", ""],
        ["", ""],
        ["2. Dicas de preenchimento:", ""],
        ["   • Valores: Use ponto ou vírgula para decimais (ex: 1250,50 ou 1250.50)", ""],
        ["   • Responsável: Use o NOME COMPLETO do usuário cadastrado no sistema", ""],
        ["   • Nomes: O sistema aceita letras maiúsculas ou minúsculas e ignora espaços extras", ""],
        ["", ""],
        ["3. Exemplos de Valores Aceitos:", ""],
        ["   • Status: 'orcamento' ou 'Orçamento' ou 'ORÇAMENTO'", ""],
        ["   • Prioridade: 'media' ou 'Média' ou 'MEDIA'", ""],
        ["   • Categoria: 'material' ou 'Material, Serviço'", ""],
        ["", ""],
        ["4. Limites:", ""],
        ["   • Máximo de 100 pedidos por arquivo", ""],
        ["   • Remova as linhas de exemplo (em azul) antes de enviar", ""],
    ]
    
    # Add status list to instructions
    for status_code, status_name in AcquisitionRequest.STATUS_CHOICES:
        instructions.append([f"   • {status_code}: {status_name}", ""])
    
    for row_idx, (col1, col2) in enumerate(instructions, 1):
        ws_instructions.cell(row=row_idx, column=1, value=col1)
        ws_instructions.cell(row=row_idx, column=2, value=col2)
        
        if row_idx == 1:  # Title
            ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
        elif col1.startswith(("1.", "2.", "3.", "4.")):  # Section headers
            ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True)
    
    # Set column widths for instructions
    ws_instructions.column_dimensions['A'].width = 60
    ws_instructions.column_dimensions['B'].width = 20
    
    return wb

def process_import_file(file_path, current_user):
    """Processa o arquivo Excel importado e retorna lista de pedidos para criação"""
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        pedidos = []
        erros = []
        
        # Mapping for flexible text matching
        status_map = {s[1].lower(): s[0] for s in AcquisitionRequest.STATUS_CHOICES}
        status_map.update({s[0].lower(): s[0] for s in AcquisitionRequest.STATUS_CHOICES})
        
        priority_map = {p[1].lower(): p[0] for p in AcquisitionRequest.PRIORITY_CHOICES}
        priority_map.update({p[0].lower(): p[0] for p in AcquisitionRequest.PRIORITY_CHOICES})
        
        impact_map = {i[1].lower(): i[0] for i in AcquisitionRequest.IMPACT_CHOICES}
        impact_map.update({i[0].lower(): i[0] for i in AcquisitionRequest.IMPACT_CHOICES})
        
        classe_map = {
            'ensino': 'ensino', 'manutenção': 'manutencao', 'manutencao': 'manutencao',
            'manutenção ': 'manutencao', ' ensino': 'ensino'
        }
        
        # Skip header row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # Skip empty rows
                continue
            
            # Ensure row has enough columns
            row_data = list(row) + [None] * (12 - len(row))
            titulo, descricao, status, prioridade, impacto, classe, categoria, data_solicitacao, valor_estimado, valor_final, responsavel_nome, observacoes = row_data[:12]
            
            # Clean text data
            titulo = str(titulo).strip() if titulo else ""
            descricao = str(descricao).strip() if descricao else ""
            
            # Validate required fields
            if not titulo or len(titulo) < 5:
                erros.append(f"Linha {row_idx}: Título é obrigatório e deve ter pelo menos 5 caracteres")
                continue
                
            if not descricao or len(descricao) < 10:
                erros.append(f"Linha {row_idx}: Descrição é obrigatória e deve ter pelo menos 10 caracteres")
                continue
            
            # Flexible mapping for status, priority, impact, classe
            status_val = status_map.get(str(status).strip().lower()) if status else None
            if not status_val:
                erros.append(f"Linha {row_idx}: Status inválido '{status}'. Use: Orçamento, Fase de Compra, etc.")
                continue
                
            priority_val = priority_map.get(str(prioridade).strip().lower()) if prioridade else "media"
            impact_val = impact_map.get(str(impacto).strip().lower()) if impacto else "medio"
            classe_val = classe_map.get(str(classe).strip().lower()) if classe else "ensino"
            
            # Category validation
            cat_val = str(categoria).strip().lower() if categoria else "material"
            valid_cats = ['material', 'servico', 'material,servico', 'serviço', 'material,serviço']
            if cat_val not in valid_cats:
                cat_val = "material" # Default
            cat_val = cat_val.replace('serviço', 'servico')
                
            # Validate and parse date
            try:
                if data_solicitacao:
                    from datetime import datetime
                    if isinstance(data_solicitacao, str):
                        # Try common formats
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                            try:
                                data_parsed = datetime.strptime(data_solicitacao.strip(), fmt).date()
                                break
                            except ValueError:
                                continue
                        else:
                            raise ValueError
                    else:
                        data_parsed = data_solicitacao.date() if hasattr(data_solicitacao, 'date') else data_solicitacao
                else:
                    from datetime import date
                    data_parsed = date.today()
            except Exception:
                from datetime import date
                data_parsed = date.today()
            
            # Find responsible user - flexible matching
            responsible_id = None
            if responsavel_nome:
                resp_name = str(responsavel_nome).strip()
                responsible = User.query.filter(User.full_name.ilike(resp_name)).first()
                if responsible:
                    responsible_id = responsible.id
                else:
                    erros.append(f"Linha {row_idx}: Responsável '{responsavel_nome}' não encontrado.")
            
            # Parse values
            def parse_float(val):
                if val is None or val == "": return None
                try:
                    return float(str(val).replace('R$', '').replace('.', '').replace(',', '.').strip())
                except: return None

            valor_estimado_parsed = parse_float(valor_estimado)
            valor_final_parsed = parse_float(valor_final)
            
            pedidos.append({
                'titulo': titulo,
                'descricao': descricao,
                'status': status_val,
                'priority': priority_val,
                'impact': impact_val,
                'classe': classe_val,
                'categoria': cat_val,
                'data_solicitacao': data_parsed,
                'valor_estimado': valor_estimado_parsed,
                'valor_final': valor_final_parsed,
                'responsible_id': responsible_id,
                'observacoes': str(observacoes).strip() if observacoes else None,
                'linha': row_idx
            })
            
            if len(pedidos) >= 100:
                erros.append("Limite de 100 pedidos atingido.")
                break
        
        return pedidos, erros
        
    except Exception as e:
        return [], [f"Erro crítico: {str(e)}"]